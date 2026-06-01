"""phase4-doc-skills：HTML one-pager (Kami) + IB PPT (ib_master) 单元测试。

策略：
- HTML：mock LLM 返合规 Kami HTML / 各类违规 HTML → 验证 invariants 校验、抽取、metadata
- PPTX：mock LLM 返 example_data.json / 各类异常输入 →
  - 有 node 可执行 → 真跑 ``node render.mjs`` 验证 .pptx 产物（zipfile 看 slide1.xml）
  - 无 node → ``pytest.skip``，不让 CI 失败
- Legacy 回滚：``use_legacy_html_pptx=True`` 时跳回 ``test_skill_executor.py`` 覆盖的旧路径
  （在本文件里只跑一条 sanity 用例，确认 settings 开关生效）

注意：和老 ``test_skill_executor.py`` 的边界——后者覆盖 legacy / 公共工具函数；本文件
**只**覆盖新的 phase4-doc-skills 链路。
"""

from __future__ import annotations

import asyncio
import json
import shutil
import zipfile
from pathlib import Path
from typing import Any

import pytest
from app.adapters.skill import SkillError, SkillExecutor
from app.adapters.skill.llm_skill import (
    _PPT_IB_DECK_FIELDS,
    _build_strategy_deck_js,
    _check_html_one_pager_invariants,
    _extract_html_document,
    _parse_deck_json,
    _parse_ib_deck_json,
    _select_doc_variant,
    _select_pptx_variant,
)
from app.adapters.skill.python_executor import ExecResult
from app.config import Settings
from app.schemas.llm import ChatMessage, LLMResponse, LLMUsage

_FIXTURE_DATA_PATH = (
    Path(__file__).resolve().parents[2]
    / "app"
    / "adapters"
    / "skill"
    / "assets"
    / "ppt_ib_deck"
    / "example_data.json"
)


class FakeLLM:
    """同 test_skill_executor.FakeLLM；复制是为了避免 cross-file import。"""

    def __init__(self, content: str) -> None:
        self.content = content
        self.last_messages: list[ChatMessage] | None = None

    async def chat(self, messages: list[ChatMessage], **_: Any) -> LLMResponse:
        self.last_messages = list(messages)
        return LLMResponse(
            content=self.content,
            model="MiniMax-M2.7",
            finish_reason="stop",
            usage=LLMUsage(prompt_tokens=10, completion_tokens=20, total_tokens=30),
            latency_ms=12.0,
        )

    async def chat_stream(self, messages: list[ChatMessage], **_: Any):  # type: ignore[no-untyped-def]
        # 2026-05-28: skill._call_llm 改走 chat_stream；mock 单 chunk 返回 self.content。
        self.last_messages = list(messages)
        yield self.content


def _settings(tmp_path: Path, *, use_legacy_html_pptx: bool = False) -> Settings:
    """phase4-doc-skills 默认 ``use_legacy_html_pptx=False``，走 Kami / IB deck 高质量路径。"""
    return Settings(
        storage_dir=tmp_path,
        skill_executor_build_dir=tmp_path / "skill_build",
        skill_executor_timeout_s=30,
        skill_executor_max_tokens=80_000,
        use_legacy_html_pptx=use_legacy_html_pptx,
    )


# ──────────────────────────────────────────────────────────────────────────
# HTML one-pager：纯函数 _extract_html_document / _check_html_one_pager_invariants
# ──────────────────────────────────────────────────────────────────────────


def _make_valid_kami_html(extra_chars: int = 6500) -> str:
    """生成最小合规 Kami HTML：含 #f5f4ed、3+ SVG、≥ 6000 chars、无 rgba/emoji/片假名。"""
    svgs = "\n".join(
        f"<svg viewBox='0 0 100 50'><line x1='0' y1='25' x2='100' y2='25' stroke='#1B365D'/>{'x' * 30}</svg>"
        for _ in range(4)
    )
    filler = "正文段落" * extra_chars
    return (
        "<!doctype html>\n"
        "<html lang='zh'><head><meta charset='utf-8'>"
        "<style>body{background:#f5f4ed;font-family:serif;}</style></head>"
        f"<body><h1>英伟达 FY2026-FY2027 投资展望</h1>{svgs}<p>{filler}</p></body></html>"
    )


@pytest.mark.unit
def test_extract_html_document_strips_leading_prose() -> None:
    raw = "Sure, here is the HTML:\n\n<!doctype html><html><body>x</body></html>\n\nDone."
    extracted = _extract_html_document(raw)
    assert extracted.startswith("<!doctype html>")
    assert extracted.endswith("</html>")


@pytest.mark.unit
def test_extract_html_document_handles_html_tag_only() -> None:
    raw = "<html><head></head><body>hi</body></html>"
    assert _extract_html_document(raw) == raw


@pytest.mark.unit
def test_extract_html_document_returns_text_when_no_marker() -> None:
    raw = "just plain text without doctype"
    assert _extract_html_document(raw) == raw.strip()


@pytest.mark.unit
def test_html_invariants_pass_on_valid_kami() -> None:
    assert _check_html_one_pager_invariants(_make_valid_kami_html()) == []


@pytest.mark.unit
def test_html_invariants_reject_rgba() -> None:
    html = (
        _make_valid_kami_html()
        .replace("#f5f4ed", "#f5f4ed")
        .replace("background:#f5f4ed", "background:rgba(245,244,237,1)")
    )
    violations = _check_html_one_pager_invariants(html)
    assert any("rgba" in v for v in violations)


@pytest.mark.unit
def test_html_invariants_reject_missing_parchment() -> None:
    html = _make_valid_kami_html().replace("#f5f4ed", "#ffffff")
    violations = _check_html_one_pager_invariants(html)
    assert any("parchment" in v or "f5f4ed" in v for v in violations)


@pytest.mark.unit
def test_html_invariants_reject_too_short() -> None:
    html = "<!doctype html><html><body>" + "#f5f4ed " * 5 + "<svg></svg>" * 3 + "</body></html>"
    violations = _check_html_one_pager_invariants(html)
    assert any("太短" in v or "chars" in v for v in violations)


@pytest.mark.unit
def test_html_invariants_reject_too_few_svg() -> None:
    html = _make_valid_kami_html().replace("<svg", "<div", 3)
    violations = _check_html_one_pager_invariants(html)
    assert any("SVG" in v for v in violations)


@pytest.mark.unit
def test_html_invariants_reject_katakana() -> None:
    html = _make_valid_kami_html() + "<p>ワークロード</p>"
    violations = _check_html_one_pager_invariants(html)
    assert any("片假名" in v for v in violations)


@pytest.mark.unit
def test_html_invariants_reject_emoji() -> None:
    html = _make_valid_kami_html() + "<p>评级</p>"
    html = html.replace("评级", "评级\U0001f600")
    violations = _check_html_one_pager_invariants(html)
    assert any("emoji" in v for v in violations)


# ──────────────────────────────────────────────────────────────────────────
# HTML one-pager：端到端（mock LLM）
# ──────────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_one_pager_happy_path(tmp_path: Path) -> None:
    skill = SkillExecutor(_settings(tmp_path))
    llm = FakeLLM(_make_valid_kami_html())
    art = await skill.generate(llm=llm, artifact_type="html", brief="英伟达 FY2026-FY2027 投资展望")
    assert art.artifact_type == "html"
    assert art.file_path.endswith(".html")
    saved = Path(art.file_path).read_text(encoding="utf-8")
    assert saved.startswith("<!doctype html>")
    assert "#f5f4ed" in saved
    assert art.metadata["skill_variant"] == "kami_one_pager"
    assert int(art.metadata["svg_count"]) >= 3
    # 用了新 prompt：含 Kami 关键字 / 不再含老版 Tailwind dark 关键字
    assert llm.last_messages is not None
    assert "Kami" in llm.last_messages[0].content
    assert "Tailwind CDN" not in llm.last_messages[0].content


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_one_pager_strips_prose_around_doctype(tmp_path: Path) -> None:
    """LLM 在 HTML 前后加自然语言时，应自动抽出 <!doctype...</html>。"""
    skill = SkillExecutor(_settings(tmp_path))
    wrapped = (
        "Sure! Here is the Kami one-pager:\n\n"
        + _make_valid_kami_html()
        + "\n\nLet me know if you'd like adjustments."
    )
    llm = FakeLLM(wrapped)
    art = await skill.generate(llm=llm, artifact_type="html", brief="x")
    saved = Path(art.file_path).read_text(encoding="utf-8")
    assert saved.startswith("<!doctype html>")
    assert "Let me know" not in saved


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_one_pager_rgba_invariant_falls_back_to_legacy(tmp_path: Path) -> None:
    """用户 2026-05-28：one-pager invariant 违反不再 400，自动降级 legacy 拿产物。

    rgba 违反 → SkillError → catch → ``_generate_via_default_pipeline`` → 用同一
    LLM 输出走 legacy 路径写盘，返回带 ``legacy_pipeline=true`` 的 artifact。
    """
    skill = SkillExecutor(_settings(tmp_path))
    bad_html = _make_valid_kami_html().replace(
        "background:#f5f4ed", "background:rgba(245,244,237,1)"
    )
    llm = FakeLLM(bad_html)
    art = await skill.generate(llm=llm, artifact_type="html", brief="x")
    assert art.artifact_type == "html"
    assert art.metadata.get("legacy_pipeline") == "true"
    saved = Path(art.file_path).read_text(encoding="utf-8")
    # legacy pipeline 不做 invariant 校验，保留原始（含 rgba）HTML
    assert "rgba(245,244,237,1)" in saved


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_one_pager_katakana_falls_back_to_legacy(tmp_path: Path) -> None:
    """日文片假名同样降级 legacy 而不是 400。"""
    skill = SkillExecutor(_settings(tmp_path))
    bad_html = _make_valid_kami_html().replace("</body>", "<p>ワークロード</p></body>")
    llm = FakeLLM(bad_html)
    art = await skill.generate(llm=llm, artifact_type="html", brief="x")
    assert art.artifact_type == "html"
    assert art.metadata.get("legacy_pipeline") == "true"


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_one_pager_invariants_unit_function() -> None:
    """``_check_html_one_pager_invariants`` 仍然作为独立函数报违规，
    供 legacy fallback 之外的诊断 / 测试使用。"""
    rgba_html = _make_valid_kami_html().replace(
        "background:#f5f4ed", "background:rgba(245,244,237,1)"
    )
    violations = _check_html_one_pager_invariants(rgba_html)
    assert any("rgba" in v for v in violations)


@pytest.mark.asyncio
@pytest.mark.unit
async def test_html_legacy_fallback_uses_old_prompt(tmp_path: Path) -> None:
    """``use_legacy_html_pptx=True`` 走旧 Tailwind dark theme 流水线。"""
    skill = SkillExecutor(_settings(tmp_path, use_legacy_html_pptx=True))
    html = (
        "<!DOCTYPE html><html><head>"
        "<script src='https://cdn.tailwindcss.com'></script></head>"
        "<body class='bg-slate-900 text-white'>" + "<p>" + ("x" * 2000) + "</p></body></html>"
    )
    llm = FakeLLM(html)
    art = await skill.generate(llm=llm, artifact_type="html", brief="legacy fallback")
    assert art.metadata.get("legacy_pipeline") == "true"
    assert llm.last_messages is not None
    assert "Tailwind CDN" in llm.last_messages[0].content


# ──────────────────────────────────────────────────────────────────────────
# IB PPT：JSON 解析（_parse_ib_deck_json 纯函数）
# ──────────────────────────────────────────────────────────────────────────


def _load_example_data() -> dict[str, str]:
    raw = json.loads(_FIXTURE_DATA_PATH.read_text(encoding="utf-8"))
    assert isinstance(raw, dict)
    return {str(k): str(v) for k, v in raw.items()}


@pytest.mark.unit
def test_pptx_variant_routes_solution_pitch_away_from_ib_template() -> None:
    brief = "帮我做一个褐蚁去 pitch 河南高校的 PPT，要教育方面的一体化解决方案，包括硬件和软件"
    assert _select_pptx_variant(brief) == "strategy"


@pytest.mark.unit
def test_pptx_variant_keeps_explicit_investment_deck_on_ib_template() -> None:
    brief = "生成英伟达 2025 投资展望 PPT，包含估值、目标价和上行空间"
    assert _select_pptx_variant(brief) == "ib"


@pytest.mark.asyncio
@pytest.mark.unit
async def test_solution_pitch_uses_strategy_json_template(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """通用方案 PPT 走 JSON → 固定模板路径；LLM 内含引号也不应破坏生成。"""
    skill = SkillExecutor(_settings(tmp_path))
    # LLM 返回 deck JSON，且内容里故意带英文引号（旧直写 JS 路径会语法崩）。
    deck = {
        "title": '从"奢侈品"走向普惠',
        "subtitle": "河南高校教育一体化",
        "footer": "来源：EchoDesk",
        "slides": [
            {"title": '痛点"分析"', "bullets": ['成本从"高"到低', "效率提升"]},
            {"title": "实施", "table": {"headers": ["阶段", "目标"], "rows": [["试点", "验证"]]}},
        ],
    }
    llm = FakeLLM(json.dumps(deck, ensure_ascii=False))

    captured: dict[str, str] = {}

    async def fake_exec(kind: str, code: str, build_dir: Path, ext: str) -> ExecResult:
        assert kind == "pptx"
        assert "const DECK =" in code  # 走了固定模板而非 LLM 直写
        captured["code"] = code
        output_path = build_dir / f"output.{ext}"
        output_path.write_bytes(b"fake pptx bytes")
        return ExecResult(success=True, output_path=output_path, stderr="", elapsed_s=0.01)

    monkeypatch.setattr(skill, "_exec_for_kind", fake_exec)
    art = await skill.generate(
        llm=llm,
        artifact_type="pptx",
        brief="褐蚁 pitch 河南高校，教育一体化解决方案，包括硬件和软件",
    )

    assert art.metadata["skill_variant"] == "strategy_pitch_deck"
    assert "legacy_pipeline" not in art.metadata
    # 用了 JSON 输出 prompt（含 schema 字段说明），不是直写 JS prompt
    assert llm.last_messages is not None
    system_prompt = llm.last_messages[0].content
    assert "结构化 JSON" in system_prompt
    assert '"slides"' in system_prompt


@pytest.mark.unit
def test_strategy_deck_builder_survives_adversarial_quotes() -> None:
    """复现真链路 bug：内容含未转义引号/反斜杠/换行，生成的 JS 仍是合法字面量。"""
    data = {
        "title": '从"奢侈品"走向\\普惠',
        "slides": [{"title": "x", "bullets": ['含"引号"', "第一行\n第二行"]}],
    }
    js = _build_strategy_deck_js(data)
    assert "const DECK =" in js
    # 提取注入的 JSON 必须能被 json.loads 还原（= 合法 JSON = 合法 JS 字面量）
    injected = js.split("const DECK = ", 1)[1].split(";\n", 1)[0]
    assert json.loads(injected)["title"] == '从"奢侈品"走向\\普惠'


class _SeqFakeLLM:
    """按调用次序返回不同 content 的 LLM（测试修复重试：先坏后好）。"""

    def __init__(self, contents: list[str]) -> None:
        self._contents = contents
        self._idx = 0
        self.last_messages: list[ChatMessage] | None = None
        self.call_count = 0

    async def chat(self, messages: list[ChatMessage], **_: Any) -> LLMResponse:
        raise NotImplementedError

    async def chat_stream(self, messages: list[ChatMessage], **_: Any):  # type: ignore[no-untyped-def]
        self.last_messages = list(messages)
        content = self._contents[min(self._idx, len(self._contents) - 1)]
        self._idx += 1
        self.call_count += 1
        yield content


@pytest.mark.asyncio
@pytest.mark.unit
async def test_xlsx_repair_retry_recovers_from_bad_code(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """LLM 直写 xlsx 代码首次执行失败（如 openpyxl ws.name 误用），
    把报错回喂修复一次后应成功，且 metadata 记录 repair_attempts=1。"""
    skill = SkillExecutor(_settings(tmp_path))
    llm = _SeqFakeLLM(
        [
            "import openpyxl\nwb=openpyxl.Workbook()\n# BROKEN uses ws.name\n",
            "import openpyxl\nwb=openpyxl.Workbook()\n# FIXED uses ws.title\n",
        ]
    )

    calls = {"n": 0}

    async def fake_exec(kind: str, code: str, build_dir: Path, ext: str) -> ExecResult:
        calls["n"] += 1
        if "BROKEN" in code:
            return ExecResult(
                success=False,
                output_path=None,
                stderr="AttributeError: 'Worksheet' object has no attribute 'name'",
                elapsed_s=0.01,
            )
        out = build_dir / f"output.{ext}"
        out.write_bytes(b"ok" * 100)
        return ExecResult(success=True, output_path=out, stderr="", elapsed_s=0.01)

    monkeypatch.setattr(skill, "_exec_for_kind", fake_exec)
    art = await skill.generate(llm=llm, artifact_type="xlsx", brief="季度营收对比表")

    assert art.metadata["repair_attempts"] == "1"
    assert calls["n"] == 2  # 初次 + 修复后各执行一次
    assert llm.call_count == 2  # 初次生成 + 修复各调一次 LLM
    # 修复调用的 user prompt 应带上报错，便于 LLM 定位
    assert llm.last_messages is not None
    assert "AttributeError" in llm.last_messages[1].content


@pytest.mark.asyncio
@pytest.mark.unit
async def test_skill_llm_stream_retries_on_zero_token_stall(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """上游 0 token 停顿（idle timeout 且 received 0 chars）应自动重连一次后成功。"""
    import app.adapters.skill.llm_skill as skill_mod

    monkeypatch.setattr(skill_mod, "_STREAM_IDLE_TIMEOUT_S", 0.05)
    skill = SkillExecutor(_settings(tmp_path))

    class _StallThenOkLLM:
        def __init__(self) -> None:
            self.calls = 0

        async def chat(self, messages: list[ChatMessage], **_: Any) -> LLMResponse:
            raise NotImplementedError

        async def chat_stream(self, messages: list[ChatMessage], **_: Any):  # type: ignore[no-untyped-def]
            self.calls += 1
            if self.calls == 1:
                # 第一次：建连成功但久不吐 token → wait_for 超时（0 chars stall）
                await asyncio.sleep(5)
                yield "unreachable"
            else:
                yield "import openpyxl\nwb=openpyxl.Workbook()\nwb.save('output.xlsx')\n"

    llm = _StallThenOkLLM()

    async def fake_exec(kind: str, code: str, build_dir: Path, ext: str) -> ExecResult:
        out = build_dir / f"output.{ext}"
        out.write_bytes(b"x" * 200)
        return ExecResult(success=True, output_path=out, stderr="", elapsed_s=0.01)

    monkeypatch.setattr(skill, "_exec_for_kind", fake_exec)
    art = await skill.generate(llm=llm, artifact_type="xlsx", brief="季度表")

    assert llm.calls == 2  # 第一次 stall → 重连第二次成功
    assert art.artifact_id.startswith("xlsx-")


@pytest.mark.unit
def test_parse_deck_json_strips_fence_and_keeps_structure() -> None:
    payload = {"title": "t", "slides": [{"title": "a", "bullets": ["b"]}]}
    raw = "```json\n" + json.dumps(payload, ensure_ascii=False) + "\n```"
    parsed = _parse_deck_json(raw)
    assert isinstance(parsed["slides"], list)
    assert parsed["slides"][0]["title"] == "a"


@pytest.mark.unit
def test_doc_variant_routes_general_away_from_finance_template() -> None:
    # 日常表格/文档 → general，不该套财务模型 / 投行研报
    assert _select_doc_variant("xlsx", "做一个班级值日表，按周排班") == "general"
    assert _select_doc_variant("xlsx", "公司物资库存清单，含数量和位置") == "general"
    assert _select_doc_variant("word", "写一份端午放假通知") == "general"
    assert _select_doc_variant("word", "整理今天的会议纪要，列出待办") == "general"


@pytest.mark.unit
def test_doc_variant_keeps_finance_template_on_finance_brief() -> None:
    assert _select_doc_variant("xlsx", "搭一个英伟达 DCF 估值财务模型") == "finance"
    assert _select_doc_variant("word", "写一份英伟达投资分析研究报告，含估值") == "finance"


@pytest.mark.unit
def test_doc_variant_empty_for_non_doc_kinds() -> None:
    assert _select_doc_variant("pptx", "随便什么") == ""
    assert _select_doc_variant("html", "随便什么") == ""


@pytest.mark.asyncio
@pytest.mark.unit
async def test_general_xlsx_uses_general_prompt_not_dcf(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """非财务 Excel 请求应走通用表格 prompt（不含 DCF/WACC 强制），并打 general 变体标签。"""
    skill = SkillExecutor(_settings(tmp_path))
    llm = FakeLLM("import openpyxl\nwb=openpyxl.Workbook()\nwb.save('output.xlsx')\n")

    async def fake_exec(kind: str, code: str, build_dir: Path, ext: str) -> ExecResult:
        out = build_dir / f"output.{ext}"
        out.write_bytes(b"x" * 200)
        return ExecResult(success=True, output_path=out, stderr="", elapsed_s=0.01)

    monkeypatch.setattr(skill, "_exec_for_kind", fake_exec)
    art = await skill.generate(llm=llm, artifact_type="xlsx", brief="做一个班级值日表，按周排班")

    assert art.metadata["skill_variant"] == "xlsx_general"
    assert llm.last_messages is not None
    sys_prompt = llm.last_messages[0].content
    # 通用自适应模板：先设计结构、不套固定财务模板、明确禁止无端 DCF
    assert "先判断" in sys_prompt
    assert "不套用任何固定财务模板" in sys_prompt
    assert "禁止" in sys_prompt and "DCF" in sys_prompt


@pytest.mark.unit
def test_parse_ib_deck_json_direct() -> None:
    raw = json.dumps({"a": "1", "b": "2"})
    assert _parse_ib_deck_json(raw) == {"a": "1", "b": "2"}


@pytest.mark.unit
def test_parse_ib_deck_json_fenced() -> None:
    payload = json.dumps({"a": "1"})
    raw = f"```json\n{payload}\n```"
    assert _parse_ib_deck_json(raw) == {"a": "1"}


@pytest.mark.unit
def test_parse_ib_deck_json_with_prose() -> None:
    raw = (
        "Sure, here is the JSON you requested:\n\n"
        '{"cover_title": "T", "kpi1_value": "$1B"}\n\n'
        "Let me know if you need adjustments."
    )
    parsed = _parse_ib_deck_json(raw)
    assert parsed["cover_title"] == "T"
    assert parsed["kpi1_value"] == "$1B"


@pytest.mark.unit
def test_parse_ib_deck_json_invalid_raises() -> None:
    with pytest.raises(SkillError, match="ib_pptx JSON"):
        _parse_ib_deck_json("not a json at all")


@pytest.mark.unit
def test_parse_ib_deck_json_coerces_non_string_values() -> None:
    raw = '{"cover_title": "T", "kpi1_value": 130, "es_b1": null}'
    parsed = _parse_ib_deck_json(raw)
    assert parsed["kpi1_value"] == "130"
    assert parsed["es_b1"] == ""


# ──────────────────────────────────────────────────────────────────────────
# IB PPT：端到端（mock LLM + 真跑 node render.mjs）
# ──────────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
@pytest.mark.unit
async def test_ib_pptx_katakana_rejected(tmp_path: Path) -> None:
    """M2.7 偶发日文片假名 → 立即拒绝，不进入 render。"""
    skill = SkillExecutor(_settings(tmp_path))
    data = _load_example_data()
    data["th_b1"] = data["th_b1"] + " ワークロード"
    llm = FakeLLM(json.dumps(data, ensure_ascii=False))
    with pytest.raises(SkillError, match="片假名"):
        await skill.generate(llm=llm, artifact_type="pptx", brief="英伟达 投资展望 估值 目标价")


@pytest.mark.asyncio
@pytest.mark.unit
async def test_ib_pptx_missing_fields_rejected(tmp_path: Path) -> None:
    """27 字段任一缺失 → 拒绝，不进入 render（避免 docxtemplater 报无意义错）。"""
    skill = SkillExecutor(_settings(tmp_path))
    data = _load_example_data()
    del data["cover_title"]
    del data["rec_action"]
    llm = FakeLLM(json.dumps(data, ensure_ascii=False))
    with pytest.raises(SkillError, match="缺失字段"):
        await skill.generate(llm=llm, artifact_type="pptx", brief="英伟达 投资展望 估值 目标价")


@pytest.mark.asyncio
@pytest.mark.unit
async def test_ib_pptx_invalid_json_rejected(tmp_path: Path) -> None:
    skill = SkillExecutor(_settings(tmp_path))
    llm = FakeLLM("LLM 没听懂，请重新提问。")
    with pytest.raises(SkillError, match="ib_pptx JSON"):
        await skill.generate(llm=llm, artifact_type="pptx", brief="英伟达 投资展望 估值 目标价")


@pytest.mark.asyncio
@pytest.mark.unit
@pytest.mark.skipif(shutil.which("node") is None, reason="node 未在 PATH（CI 跳过）")
async def test_ib_pptx_happy_path_with_real_render(tmp_path: Path) -> None:
    """mock LLM 返回 example_data.json → 真跑 node render.mjs → 验证 .pptx 文件。

    用 zipfile 打开产物，确认 ``ppt/slides/slide1.xml`` 存在且大小合理。
    """
    deck_node_modules = (
        Path(__file__).resolve().parents[2]
        / "app"
        / "adapters"
        / "skill"
        / "assets"
        / "ppt_ib_deck"
        / "node_modules"
    )
    if not deck_node_modules.exists():
        pytest.skip("ppt_ib_deck/node_modules 缺失；先跑 scripts/install-backend.sh 或 npm install")

    skill = SkillExecutor(_settings(tmp_path))
    data = _load_example_data()
    llm = FakeLLM(json.dumps(data, ensure_ascii=False))
    art = await skill.generate(
        llm=llm,
        artifact_type="pptx",
        brief="英伟达 FY2026-FY2027 投资展望",
    )
    assert art.artifact_type == "pptx"
    out = Path(art.file_path)
    assert out.exists()
    assert out.stat().st_size > 30_000
    with zipfile.ZipFile(out) as zf:
        names = zf.namelist()
    assert "ppt/slides/slide1.xml" in names
    assert any(n.startswith("ppt/slides/slide") for n in names)
    assert art.metadata["skill_variant"] == "ib_deck_v3"
    assert art.metadata["slide_count_hint"] == "14"
    assert art.metadata["field_count"] == str(len(_PPT_IB_DECK_FIELDS))

    data_json = (out.parent / "data.json").read_text(encoding="utf-8")
    saved = json.loads(data_json)
    assert set(saved.keys()) == set(_PPT_IB_DECK_FIELDS)


@pytest.mark.asyncio
@pytest.mark.unit
async def test_ib_pptx_node_missing_raises(tmp_path: Path) -> None:
    """node 不存在时清晰报错而非 crash。"""
    skill = SkillExecutor(
        Settings(
            storage_dir=tmp_path,
            skill_executor_build_dir=tmp_path / "skill_build",
            skill_executor_timeout_s=10,
            skill_executor_max_tokens=80_000,
            skill_node_bin="/non/existent/node-xyz",
            use_legacy_html_pptx=False,
        )
    )
    data = _load_example_data()
    llm = FakeLLM(json.dumps(data, ensure_ascii=False))
    with pytest.raises(SkillError, match=r"node|render"):
        await skill.generate(llm=llm, artifact_type="pptx", brief="x")


@pytest.mark.unit
def test_ppt_ib_deck_assets_exist() -> None:
    """phase4-doc-skills：assets/ppt_ib_deck/ 三个关键文件必须在 repo 里。

    node_modules 不强求（由 scripts/install-backend.sh 装；CI 用 ``test_ib_pptx_happy_path_with_real_render``
    在该目录缺失时 ``skip``）。
    """
    base = (
        Path(__file__).resolve().parents[2]
        / "app"
        / "adapters"
        / "skill"
        / "assets"
        / "ppt_ib_deck"
    )
    assert (base / "ib_master.pptx").is_file()
    assert (base / "render.mjs").is_file()
    assert (base / "package.json").is_file()
    assert (base / "example_data.json").is_file()
    assert (base / "schema.md").is_file()
